from copyreg import dispatch_table
import time
import os

# ------------------------------------------------------------------------------------------
print("***** Put the interpreter in UTC, to make sure no TZ issues")
os.environ["TZ"] = "UTC"
time.tzset()

from icecream import ic

import datetime
import pytz

import pickle as pkl

utc_timezone = pytz.timezone("UTC")

ic.configureOutput(prefix="", outputFunction=print)

from openpyxl import load_workbook

dict_all_data = {}

print("start loading Excel sheet (this may take a few seconds)")
wb = load_workbook("./Svalbard_Ice_Tracker_Data_2017.xlsx")

for crrt_instrument in wb.sheetnames:
    ic(crrt_instrument)

    # quality check the header line
    assert wb[crrt_instrument]["A1"].value == "ReadingDate", f"expected 'ReadingDate', got {wb[crrt_instrument]['A1'].value}"
    assert wb[crrt_instrument]["H1"].value == "Latitude", f"expected 'Latitude', got {wb[crrt_instrument]['H1'].value}"
    assert wb[crrt_instrument]["I1"].value == "Longitude", f"expected 'Longitude', got {wb[crrt_instrument]['I1'].value}"

    dict_all_data[crrt_instrument] = {}

    # fill the entries
    crrt_line = 2
    while wb[crrt_instrument][f"A{crrt_line}"].value is not None:
        crrt_datetime = wb[crrt_instrument][f"A{crrt_line}"].value
        crrt_lat = wb[crrt_instrument][f"H{crrt_line}"].value
        crrt_lon = wb[crrt_instrument][f"I{crrt_line}"].value
        dict_all_data[crrt_instrument][crrt_datetime] = {}
        dict_all_data[crrt_instrument][crrt_datetime]["lat"] = crrt_lat
        dict_all_data[crrt_instrument][crrt_datetime]["lon"] = crrt_lon
        crrt_line += 1


print("dump")

# dump with pickle
with open("data_from_drifter_Ian.pkl", "bw") as fh:
    pkl.dump(dict_all_data, fh)
